import datetime, os, re, sys, tkinter as tk
import configparser as cp, cv2, numpy as np, openpyxl, pandas as pd, pytesseract
from copy import copy
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import messagebox, ttk
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
from tkinter.simpledialog import askinteger

class TextRedirector(object):
    def __init__(self, tb, tag= 'stdout'):
        self.tb = tb
        self.tag = tag

    def write(self, str):
        self.tb.configure(state= 'normal')
        self.tb.insert('end', str, (self.tag,))
        self.tb.update_idletasks()
        self.tb.see(tk.END)
        self.tb.configure(state= 'disabled')

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath('.')

    return os.path.join(base_path, relative_path)

def generateDropLog(tb, bosses, xltf, kwf, isVerbose, option):
    tb.configure(state= 'normal')
    tb.delete(1.0, tk.END)
    tb.configure(state= 'disabled')

    if option == 'updateExistingSheet':
        if messagebox.askokcancel(icon= 'warning', title= 'Caution', message= 'Update mode is currently selected, which EDITS a loaded Excel sheet instead of generating a new one.\n\nALL CHANGES ARE IRREVERSIBLE.\nONLY PROCEED IF YOU KNOW WHAT YOU ARE DOING!\n\nPress OK to proceed, or Cancel to abort.'):
            xltfile = askopenfilename(title= 'Select Excel Workbook to update', initialdir= os.getcwd(), filetypes= [('Excel Workbook', '.xlsx'), ('Excel 97- Excel 2003 Workbook', '.xls')])
            wb = openpyxl.load_workbook(xltfile)
            print('Excel template loaded:', os.path.basename(xltfile), '\n')
            ws = wb[wb.sheetnames[askinteger(title= 'Select sheet', prompt= 'Enter the number of the sheet to edit (0 to {}):'.format(len(wb.sheetnames) - 1))]]
            df = createDataFrame(wb, ws)
        else:
            return
    else:
        xltfile = xltf.get()
        wb = openpyxl.load_workbook(xltfile)
        ws = wb[wb.sheetnames[0]]
        df = createDataFrame(wb, ws)
        time = datetime.datetime.now()
        ws.title = time.strftime('%d %b %Y %H-%M')
    kwfile = kwf.get()
    pytesseract.pytesseract.tesseract_cmd = resource_path('tess\\tesseract.exe')
    
    keywords = []
    images = []
    
    try:
        images.extend(askopenfilenames(title= 'Select image files to process', initialdir= os.getcwd(), filetypes= [('Image files', '.jpg .jpeg .png .bmp .tiff')]))
        if len(images) == 0:
            raise FileNotFoundError('No image file(s) selected!')
    except FileNotFoundError as e:
        print(e, '\nOperation aborted.')
        return
    with open(kwfile) as kw:
        for line in kw:
            keywords.append(line.rstrip('\n'))
    
    for img_path in images:
        drop_list = {}
        print('Reading in', img_name := os.path.basename(img_path), '... ', end= '')
        img = cv2.imread(img_path)
        print('Done!')
        
        # Image pre-processing: Greyscaling, Otsu's thresholding, resizing, Otsu's thresholding again, Gaussian blurring
        print('Processing', img_name, '... ', end= '')
        inv = np.invert(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY))
        ot, ot_result = (otsu := lambda input : cv2.threshold(input, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU))(inv)
        scale_factor = 4
        width = int(img.shape[1] * scale_factor)
        height = int(img.shape[0] * scale_factor)
        rsz = cv2.resize(ot_result, (width, height), interpolation = cv2.INTER_AREA)
        ot_2, ot_result_2 = otsu(rsz)
        blur = cv2.GaussianBlur(ot_result_2, (5, 5), 0)
        print('Done!')
        
        # Pass processed image to Tesseract OCR
        print('Detecting text in', img_name, '... ', end= '')
        output = pytesseract.image_to_string(blur, lang= 'eng',config= '--psm 4 --oem 1')
        rep1 = re.sub('v¥|¥V|WY|YY|VV|VY|vY|WV|vv|“W|“Y|“v', 'W', output) # This regex corrects the problematic capital W in the source image
        rep2 = re.sub('vy', 'w', rep1)
        drops = rep2.split('\n')
        drops = list(filter(None, drops))
        print('Done!')
        
        if isVerbose is True:
            print('Tesseract output:', dash := '----------------------------------------', *drops, dash, sep= '\n')
        
        for i, item in enumerate(df.Item.values.tolist()):
            if item == None:
                break
            drop_list[item] = 0
        
        for b in bosses:
            if b.casefold() in img_name.casefold():
                print('Recording drops from', b, '... ', end= '')
                
                # Retrieve drop quantities from Tesseract output
                for kw in keywords:
                    for dr in drops:
                        if f'{kw}' in dr:
                            qnt = re.search('[x]{1}\d+', dr)
                            quantity = int(qnt.group(0).split('x')[1])
                            for key, val in drop_list.items():
                                if f'{kw}' in key: drop_list[key] = val + quantity
                
                # Record drop quantities in DataFrame
                for row in range(len(drop_list)):
                    for key, val in drop_list.items():
                        if f'{key}' in df.at[row + 1, 'Item']:
                            df.at[row + 1, b] = val + df.at[row + 1, b] if (df.at[row + 1, b] is not None) else val
                print('Done!\n')
    '''
    if isVerbose is True:
        with pd.option_context('expand_frame_repr', False, 'display.max_rows', None, 'display.max_columns', None, 'display.width', 0, 'display.max_colwidth', 0):
            print(df, end= '\n\n')
    '''
    # Write DataFrame contents to Excel sheet
    rows = dataframe_to_rows(df, index= False)

    if option == 'appendSheet':
        # Default behaviour, creates a new sheet and appends it to an existing Workbook
        wb2 = openpyxl.load_workbook(xl := askopenfilename(title= 'Select Excel Workbook to write new sheet to', initialdir= os.getcwd(), filetypes= [('Excel Workbook', '.xlsx'), ('Excel 97- Excel 2003 Workbook', '.xls')]))
        ws2 = wb2.create_sheet(time.strftime('%d %b %Y %H-%M')) # Sheets cannot be copied wholesale between two different Workbooks
        ws2.freeze_panes = 'F1'
        for r, row in enumerate(rows, 1):
            for c, value in enumerate(row, 1):
                ws2.cell(row= r, column= c, value= value).alignment = Alignment(horizontal= 'center', wrap_text= True)
                # Copy formatting from Excel template to new sheet
                if ws.cell(row= r, column= c, value= value).has_style:
                    ws2.cell(row= r, column= c, value= value).font = copy(ws.cell(row= r, column= c, value= value).font)
                    ws2.cell(row= r, column= c, value= value).fill = copy(ws.cell(row= r, column= c, value= value).fill)
                    ws2.cell(row= r, column= c, value= value).number_format = copy(ws.cell(row= r, column= c, value= value).number_format)
        for i, rd in ws.row_dimensions.items():
            ws2.row_dimensions[i] = copy(rd)
        for i, cd in ws.column_dimensions.items():
            ws2.column_dimensions[i] = copy(cd)
        wb2.template = False
        print('Writing extracted drop data to', os.path.basename(xl),'... ', end= '')
        wb2.save(xl)
        print('Done!')
    
    elif option == 'createNewWorkbook':
        for r, row in enumerate(rows, 1):
            for c, value in enumerate(row, 1):
                ws.cell(row= r, column= c, value= value).alignment = Alignment(horizontal= 'center', wrap_text= True)
        wb.template = False
        print('Writing extracted drop data to new Workbook:', (os.path.basename(xl := asksaveasfilename(title= 'Create new Excel Workbook', filetypes= [('Excel Workbook', '.xlsx'), ('Excel 97- Excel 2003 Workbook', '.xls')], defaultextension= '.xlsx'))), '... ', end= '')
        wb.save(xl)
        print('Done!')
    
    elif option == 'updateExistingSheet':
        for r, row in enumerate(rows, 1):
            for c, value in enumerate(row, 1):
                ws.cell(row= r, column= c, value= value).alignment = Alignment(horizontal= 'center', wrap_text= True)
        wb.template = False
        print('Updating', os.path.basename(xltfile),'... ', end= '')
        wb.save(xltfile)
        print('Done!')
  
# Create a Pandas DataFrame using an Excel template  
def createDataFrame(wb, ws):
    df = pd.DataFrame(ws.values)
    columnNames = df.iloc[0]
    df = df[1:]
    df.columns = columnNames
    return df

def setBoolean(cfg, bool, boolname):
    cfg.set('Booleans', boolname, bool)
    with open('dlconfig.ini', 'w') as cfgfile:
        cfg.write(cfgfile)

def readFile(cfg, label, filevar, filename, Title, types):
    old_filename = filename.get()
    filename.set(askopenfilename(title= Title, initialdir= os.getcwd(), filetypes= types))
    if filename.get() == '' or filename == None:
        filename.set(old_filename)
        return
    else:
        cfg.set('Files', filevar, filename.get())
        label.config(text= filename.get())
        with open('dlconfig.ini', 'w') as cfgfile:
            cfg.write(cfgfile)

def main():
    version = '1.3'
    window = tk.Tk()
    window.title('Drop Logging Tool v' + version)
    window.minsize(971, 600)
    window.resizable(True, True)
    frame1 = tk.Frame(master= window)
    frame2 = tk.Frame(master= window)
    frame3 = tk.Frame(master= window)
    options = {'Append': 'appendSheet', 'Create': 'createNewWorkbook', 'Update': 'updateExistingSheet'}
    option = tk.StringVar(window, value= options['Append'])
    isVerbose = tk.BooleanVar(window)
    xltfile = tk.StringVar(window, value= 'Not selected')
    kwfile = tk.StringVar(window, value= 'Not selected')
    bosslist = []
    cfg = cp.ConfigParser()
    
    try:
        cfg.read('dlconfig.ini')
        files = cfg['Files']
        booleans = cfg['Booleans']
        bosses = cfg['Bosses']
        xltfile.set(files['xltfile'])
        kwfile.set(files['kwfile'])
        isVerbose.set(booleans.getboolean('isVerbose'))
        bosslist = bosses['bosses'].split(', ')
    except (FileNotFoundError, KeyError) as e:
        cfg.add_section('Files')
        cfg.add_section('Booleans')
        cfg.add_section('Bosses')
        cfg.set('Files', 'xltfile', '')
        cfg.set('Files', 'kwfile', '')
        cfg.set('Booleans', 'isVerbose', 'False')
        cfg.set('Bosses', 'bosses', 'Lotus, Damien, Lucid, Will, Divine King Slime, Dusk, Djunkel, Heretic Hilla, Black Mage, Seren, Kalos, Kaling')
        bosslist = ['Lotus', 'Damien', 'Lucid', 'Will', 'Divine King Slime', 'Dusk', 'Djunkel', 'Heretic Hilla', 'Black Mage', 'Seren', 'Kalos', 'Kaling']
        with open('dlconfig.ini', 'w') as cfgfile:
            cfg.write(cfgfile)
    
    loadExcelButton = ttk.Button(master= frame1, text= 'Excel template', width= 20, command= lambda: readFile(cfg, showCurrentExcelLabel, 'xltfile', xltfile, 'Select Excel template', [('Template', '.xltx'), ('Template (code)', '.xltm'), ('Excel Workbook', '.xlsx'), ('Excel 97- Excel 2003 Workbook', '.xls')]))
    loadKeywordsButton = ttk.Button(master= frame1, text= 'Keywords', width= 20, command= lambda: readFile(cfg, showCurrentKeywordsLabel, 'kwfile', kwfile, 'Select keywords text file', [('Text file', '.txt'), ('All files', '.*')]))

    showCurrentExcelLabel = ttk.Label(master= frame1, text= xltfile.get())
    showCurrentKeywordsLabel = ttk.Label(master= frame1, text= kwfile.get())

    generateDropLogButton = ttk.Button(master= frame2, text= 'Generate Excel sheet', width= 30, command= lambda: generateDropLog(textbox, bosslist, xltfile, kwfile, isVerbose.get(), option.get()))
    isVerboseButton = ttk.Checkbutton(master= frame2, text= 'Verbose', variable= isVerbose, command= lambda: setBoolean(cfg, str(isVerbose.get()), 'isVerbose'))
    appendSheetButton = ttk.Radiobutton(master= frame2, text= 'Append sheet to existing Workbook', variable= option, value= options['Append'], command= lambda: loadExcelButton.config(state= 'enabled'))
    createNewWorkbookButton = ttk.Radiobutton(master= frame2, text= 'Create new Workbook', variable= option, value= options['Create'], command= lambda: loadExcelButton.config(state= 'enabled'))
    updateExistingSheetButton = ttk.Radiobutton(master= frame2, text= 'Update existing sheet', variable= option, value= options['Update'], command= lambda: loadExcelButton.config(state= 'disabled'))
    
    textbox = tk.Text(master= frame3, font= ('Consolas', 10), wrap= tk.NONE, state= 'disabled')
    vert_scrollbar = ttk.Scrollbar(master= frame3, orient= 'vertical', command= textbox.yview)
    hori_scrollbar = ttk.Scrollbar(master= frame3, orient= 'horizontal', command= textbox.xview)
    
    frame1.pack()
    frame2.pack()
    frame3.pack(side= 'bottom', fill= 'both', expand= True)

    frame1.columnconfigure(1, weight= 1)
    frame1.columnconfigure(2, weight= 1, minsize= 400)

    loadExcelButton.grid(row= 1, column= 1, padx= (0, 10), pady= (15, 5))
    showCurrentExcelLabel.grid(row= 1, column= 2, sticky= 'w', pady= (15, 5))

    loadKeywordsButton.grid(row= 2, column= 1, padx= (0, 10), pady= 5)
    showCurrentKeywordsLabel.grid(row= 2, column= 2, sticky= 'w', pady= 5)

    generateDropLogButton.grid(row= 1, column= 1, columnspan= 2, pady= (10, 5))
    isVerboseButton.grid(row= 1, column= 2, columnspan= 2, pady= (10, 5))
    appendSheetButton.grid(row= 2, column= 1, padx= (0, 10), pady= (0, 10))
    createNewWorkbookButton.grid(row= 2, column= 2, padx= 10, pady= (0, 10))
    updateExistingSheetButton.grid(row= 2, column= 3, padx= 10, pady= (0, 10))
    
    frame3.rowconfigure(0, weight= 1)
    frame3.columnconfigure(0, weight= 1)
    textbox.grid(row= 0, column= 0, sticky= 'nsew')
    vert_scrollbar.grid(row= 0, column= 1, sticky= 'ns')
    textbox['yscrollcommand'] = vert_scrollbar.set
    hori_scrollbar.grid(row= 1, column= 0, sticky= 'ew')
    textbox['xscrollcommand'] = hori_scrollbar.set
    sys.stdout = TextRedirector(textbox, 'stdout')
    sys.stderr = TextRedirector(textbox, 'stderr')
    
    print('Drop Logging Tool version', version)
    print('\n--------------------\n')
    print('Ready')
    window.mainloop()

if __name__ == '__main__':
    main()